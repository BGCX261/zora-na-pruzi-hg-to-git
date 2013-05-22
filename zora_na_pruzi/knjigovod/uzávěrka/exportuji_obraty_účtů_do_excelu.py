#!/usr/bin/env python3
# Copyright (c) 2012 медвед медведович медведев <gazda@xtrip.net>

'''
Hen je program, který exportuje obraty účtů do excelu
používám dva excelovské soubory
do jednoho importuji z databáze obraty účtů
v druhém mám výkazy
je to z toho důvodů, že openpyxl zruší formáty buněk, když se uloží
'''

#TODO: Až bude openpyxl zachovávat formátování buněk, tak budu importovat přímo do excelovského souboru výkazů

import os
from openpyxl import Workbook

from pruga.knjigovod.nastavení.databáze import davaj_připojení,   JMÉNO_DATABÁZE

def __sestavím_cestu_k_souboru(rok,  jméno_druhu_souboru):
    jméno_souboru = 'excel/závěrka-{rok}-{druh_souboru}.xlsx'.format(rok = rok,  druh_souboru = jméno_druhu_souboru)
    cesta_k_souboru = os.path.join(os.path.dirname(__file__),  jméno_souboru)
    return cesta_k_souboru
    
def sestavím_cestu_k_souboru_účtů(rok):
    return __sestavím_cestu_k_souboru(rok, 'obraty_účtů')
    
def sestavím_cestu_k_souboru_výkazů(rok):
    return __sestavím_cestu_k_souboru(rok, 'výkazy')

def zapíšu_obraty_účtů_do_excelu(rok = 2011):
    db = davaj_připojení(JMÉNO_DATABÁZE)
    
    cesta_k_souboru = sestavím_cestu_k_souboru_účtů(rok)
    
    sešit = Workbook()
    
    print('exportuji obraty syntetických účtů')
    
    list = sešit.get_active_sheet()
    list.title = 'syntetické_účty'
    
    sql = 'SELECT "rok", "číslo_účtu", "jméno_účtu", "obrat_má_dáti", "obrat_dal", "konečný_zůstatek" FROM "uzávěrka_obratu_účtů_souhrn_syntetických_účtů" WHERE "rok" = {rok} ORDER BY "číslo_účtu"'.format(rok = rok)
    obraty =  db.prepare(sql)
    
    vypíšu_do_excelu(list,  obraty)

    print('exportuji obraty analytických účtů')
    
    list = sešit.create_sheet()
    list.title = 'analytické_účty'
    
    sql = 'SELECT "rok", ("číslo_účtu")."syntetický" AS "syntetický účet", ("číslo_účtu")."analytický" AS "analytický účet", "jméno_účtu", "obrat_má_dáti", "obrat_dal", "konečný_zůstatek" FROM "pohled_uzávěrka_obraty_účtů" WHERE "rok" = {rok} ORDER BY "číslo_účtu"'.format(rok = rok)
    obraty =  db.prepare(sql)
    
    vypíšu_do_excelu(list,  obraty)
    
    sešit.save(cesta_k_souboru)
    
    print('uloženo v {}'.format(cesta_k_souboru))
    
   

def vypíšu_do_excelu(list,  data):
    for sloupec,  nadpis in enumerate(data.column_names):
        list.cell(row = 0,  column = sloupec).value = nadpis
    
    for číslo_řádku,  řádek in enumerate(data()):
        for číslo_sloupce,  obsah in enumerate(řádek):
            list.cell(row = číslo_řádku + 1,  column = číslo_sloupce).value = obsah


if __name__ == '__main__':

    print(__doc__)
    
    import argparse

    #  nejdříve si parser vytvořím
    parser = argparse.ArgumentParser(description='Uzávěrka, spočtu obraty účtů.')

#   a pak mu nastavím jaké příkazy a parametry má přijímat
#    parser.add_argument('--version', '-v',  action='version', version='%(prog)s, pruga verze: Пруга {}'.format(__version__))

    parser.add_argument('rok',  type=int)
    args = parser.parse_args()
    
    zapíšu_obraty_účtů_do_excelu(args.rok)
    


