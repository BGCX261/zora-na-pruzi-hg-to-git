#!/usr/bin/env python3
# Copyright (c) 2012 медвед медведович медведев <gazda@xtrip.net>

'''
Hen je program, který ...
'''

import os
from openpyxl import load_workbook

def načtu_body_z_excelu(cesta_k_souboru):
    '''
    
    '''
#    cesta_k_souboru = 'transform GP.xlsx'
    
    sešit = load_workbook(cesta_k_souboru)
#    print(sešit)
    list = sešit.get_sheet_by_name("zJTSK")
#    print(sešit.get_sheet_names())
#    print(list)
    
    body = {}
    for řádek in range(2,  1000):
        číslo_bodu = list.cell(row = řádek, column = 0).value
        if číslo_bodu is None:
            break
        poledník = list.cell(row = řádek, column = 10).value
        rovnoběžka =  list.cell(row = řádek, column = 9).value
#        print(číslo_bodu, poledník, rovnoběžka)
        body[číslo_bodu] = [rovnoběžka,  poledník]
      
    return body
    
def vytvořím_html_smapu():

#    body = načtu_body_z_excelu('transform.xlsx')
#    
#
    js_kód = {}
#    js_kód['body'] = davaj_js_kód_bodů(body)
    
    body = načtu_body_z_excelu('transform GP.xlsx')
    
    pozemky = davaj_souřadnice_pozemků(body)
    js = '''
    var options_pozemku = {
            color: "#f00",
            width: 3
        };
    '''
    js_polygony = []
    js_popisky = []
    js_polygony.append(js)
    for číslo_pozemku,  souřadnice_pozemku in pozemky.items():
        js_polygony.append(davaj_js_kód_pozemku(číslo_pozemku,  souřadnice_pozemku))
        js_popisky.append(davaj_js_kód_popisu_pozemku(číslo_pozemku,  souřadnice_pozemku))
    
    js_kód['polygony'] = '\n'.join(js_polygony)
    js_kód['body'] = '\n'.join(js_popisky)
    
#    hen popis pozemků
    
    
    
    with open(file = 'html/mapa.html',  encoding = 'utf8',  mode = 'r') as html_šablona:
        html = html_šablona.read()

    html = html.format(**js_kód)
    
    print(html)

def davaj_js_kód_bodů(body):
    
    js_kód = []
    
    for číslo_bodu,  souřadnice in body.items():
        rovnoběžka,  poledník = souřadnice
        js = '''
        var bod = SMap.Coords.fromWGS84({poledník}, {rovnoběžka});
        var element = JAK.cel('p', 'popisek');
        var txt = JAK.ctext({číslo_bodu});
        element.appendChild(txt);
        options = {{
            title: "bod {číslo_bodu}",
            url: element	 					
        }};
        var marker = new SMap.Marker(bod, "bod_{číslo_bodu}", options);
        layer_bodu.addMarker(marker);
        '''
        js_kód.append(js.format(poledník = poledník,  rovnoběžka = rovnoběžka,  číslo_bodu = číslo_bodu))    
    
    return '\n'.join(js_kód)

def davaj_js_kód_popisu_pozemku(číslo_pozemku,  souřadnice_pozemku):
    x = 0
    y = 0
    počet = len(souřadnice_pozemku)
    for souřadnice in souřadnice_pozemku:
        rovnoběžka,  poledník = souřadnice
        x = x + rovnoběžka
        y = y + poledník
        
    x = x/počet
    y = y/počet
     
    js = '''
        var bod = SMap.Coords.fromWGS84({poledník}, {rovnoběžka});
        var element = JAK.cel('p', 'popisek');
        var txt = JAK.ctext('{číslo_bodu}');
        element.appendChild(txt);
        options = {{
            title: "bod {číslo_bodu}",
            url: element	 					
        }};
        var marker = new SMap.Marker(bod, "bod_{číslo_bodu}", options);
        layer_bodu.addMarker(marker);
        '''
        
    return js.format(poledník = y,  rovnoběžka = x,  číslo_bodu = číslo_pozemku)

def davaj_souřadnice_pozemků(body):
    pozemky = {}
    pozemky['502_15'] = [1,  2,  3,  4,  5, 6,  7,  8,  9,  17,  18,  19,  20,  21,  22,  23,  25,  317,  'XX',  797,  397]
    pozemky['502_33'] = [838,  1,  2,  3,  16,  227]
    pozemky['502_34'] = [3,  4,  15,  14,  223,  16]
    pozemky['502_35'] = [4,  5,  13,  163,  14,  15]
    pozemky['502_36'] = [5,  6,  12,  13]
    pozemky['502_37'] = [6,  7,  8,  11,  798,  12]
    pozemky['502_38'] = [8,  9,  10,  11]
    pozemky['502_39'] = [9,  17,  27,  10]
    pozemky['502_40'] = [17,  18,  580,  131,  27]
    pozemky['502_41'] = [18,  19,  26,  126,  128,  420,  580]
    pozemky['502_42'] = [19,  20,  21,  22,  23,  29,  28,  24,  171,  312,  26]
    pozemky['504'] = [23,  25,  316,  313,  24,  28,  29]
     
    souřadnice_pozemků = {}
     
    for číslo_pozemku,  čísla_bodů in pozemky.items():
        souřadnice = []
        for číslo_bodu in čísla_bodů:
            bod = body[číslo_bodu]
            souřadnice.append(bod)
        souřadnice_pozemků[číslo_pozemku] = souřadnice
        
    return souřadnice_pozemků

def davaj_js_kód_pozemku(číslo_pozemku,  souřadnice_pozemku):
    
    
    js_kód = []
    js = '''
    var body_pozemku_{číslo_pozemku} = [
    '''
    js = js.format(číslo_pozemku = číslo_pozemku)
    js_kód.append(js)
    
    js_bodů = []
    for souřadnice in souřadnice_pozemku:
        rovnoběžka,  poledník = souřadnice
        js = 'SMap.Coords.fromWGS84({poledník}, {rovnoběžka})'.format(poledník = poledník,  rovnoběžka = rovnoběžka)
        js_bodů.append(js)
      
    js = ',\n'.join(js_bodů)
    js_kód.append(js)
    
    js = '''
        ];
        
        var polygon = new SMap.Geometry(SMap.GEOMETRY_POLYGON, null, body_pozemku_{číslo_pozemku}, options_pozemku);
        layer_pozemku.addGeometry(polygon);
    '''
    
    js_kód.append(js)
    
    js = '\n'.join(js_kód)
    js = js.format(číslo_pozemku = číslo_pozemku)
    return js

if __name__ == '__main__':

#    print(__doc__)

    vytvořím_html_smapu()



